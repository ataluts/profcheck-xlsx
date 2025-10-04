import sys, argparse
from datetime import datetime
from pathlib import Path
import re
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from colormath.color_objects import LabColor, sRGBColor
from colormath.color_conversions import convert_color

_module_date = datetime(2025, 10, 5)
_module_designer = "Alexander Taluts"

#Styles
style_it872_cell_size = (6, 32)    #width, height
style_it872_header_font  = Font(bold=True, size=20.0)
style_it872_header_fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
style_it872_header_alignment = Alignment(horizontal="center", vertical="center")
style_it872_value_numformat = '0.00'
style_it872_value_alignment = Alignment(horizontal="center", vertical="center")
style_it872_value_font = Font(bold=True, color="FFFFFF")
style_it872_value_fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
style_stat_header_font = Font(bold=True)
style_stat_header_fill = PatternFill()
style_stat_header_alignment = Alignment(horizontal="right", vertical="center")
style_stat_value_font = Font()
style_stat_value_fill = PatternFill()
style_stat_value_alignment = Alignment(horizontal="right", vertical="center")
style_stat_value_numformat = '0.000'
style_grade_cell_size = style_it872_cell_size
style_grade_header_font = Font(bold=True)
style_grade_header_fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
style_grade_header_alignment = Alignment(horizontal="center", vertical="center")
style_grade_value_font = Font(bold=True, color="FFFFFF")
style_grade_value_alignment = Alignment(horizontal="center", vertical="center")
style_grade_value_fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")

#dE gradations: (threshold, color above it)
dE_gradations = ((0.0, "0080FF"), (1.0, "00C000"), (2.0, "00FF00"), (3.0, "FFFF00"), (4.0, "FF8000"), (5.0, "FF0000"), (6.0, "FF00FF"))

#Class with all data about the patch
class Patch():
    def __init__(self, id, dE, error, actual, reference):
        self.id = id                #patch id: string
        self.dE = dE                #delta E value: float
        self.error = error          #per channel error?: (float, float, float)
        self.actual = actual        #actual color value (PCS): (float, float, float)
        self.reference = reference  #reference color value (PCS): (float, float, float)

#Writes column numbers of the table
def it872_header_col_num(worksheet, row):
    for col in range(1, 25):
        cell = worksheet.cell(row=row, column=col)
        cell.fill = style_it872_header_fill
        cell.font = style_it872_header_font
        cell.alignment = style_it872_header_alignment
        worksheet.column_dimensions[get_column_letter(col)].width = style_it872_cell_size[0]
        worksheet.row_dimensions[row].height = style_it872_cell_size[1]
        if 1 < col < 24:
            cell.value = col - 1

#Writes row letters of the table
def it872_header_row_letter(worksheet, col):
    for row in range(2, 14):
        cell = worksheet.cell(row=row, column=col)
        cell.fill = style_it872_header_fill
        cell.font = style_it872_header_font
        cell.alignment = style_it872_header_alignment
        cell.value = chr(row - 2 + ord('A'))
        worksheet.row_dimensions[row].height = style_it872_cell_size[1]
    worksheet.row_dimensions[15].height = style_it872_cell_size[1] * 2

#Converts patch id to cell address
def patch_to_cell(id):
    #split patch address into letter and number parts
    match = re.match(r"([A-Z]+)(\d+)", id, re.I)
    if match:
        patch_letter, patch_number = match.groups()
        patch_letter = patch_letter.upper()
        patch_number = int(patch_number)
    else:
        raise ValueError(f"Wrong patch address: {id}")
    #convert patch address to cell address
    if patch_letter == "GS":
        cell_row = 15
        cell_col = patch_number + 1
    else:
        cell_row = ord(patch_letter) - ord('A') + 2
        cell_col = patch_number + 1
    return (cell_row, cell_col)

#Parses triplet values into tuple
def parse_triplet(s):
    parts = s.strip().split()
    if len(parts) != 3:
        raise ValueError(f"Wrong string format: {s}")
    return tuple(float(x) for x in parts)

#Converts PCS color to RGB
def pcs_to_rgb(pcs_values):
    l, a, b = pcs_values
    lab = LabColor(l, a, b)
    rgb = convert_color(lab, sRGBColor)
    return tuple(int(max(0, min(255, round(c * 255)))) for c in (rgb.rgb_r, rgb.rgb_g, rgb.rgb_b))

#Reads profcheck report into patch array
def read_report(address):
    #Regex to parse report lines
    re_patch_line = re.compile(
        r"\[(?P<dE>[0-9.]+)\]\s+"
        r"(?P<patch_id>([A-Z]+)(\d+)):\s+"
        r"(?P<error>[0-9.eE+-]+\s+[0-9.eE+-]+\s+[0-9.eE+-]+)\s+->\s+"
        r"(?P<color_act>[0-9.eE+-]+\s+[0-9.eE+-]+\s+[0-9.eE+-]+)\s+should be\s+"
        r"(?P<color_ref>[0-9.eE+-]+\s+[0-9.eE+-]+\s+[0-9.eE+-]+)"
    )
    data = []
    with open(address, "r") as f:
        for line in f:
            m = re_patch_line.match(line)
            if not m:
                match = re.search(r"max\.\s*=\s*([0-9.]+),\s*avg\.\s*=\s*([0-9.]+),\s*RMS\s*=\s*([0-9.]+)", line)
                if match:
                    stat_dE_max = float(match.group(1))
                    stat_dE_avg = float(match.group(2))
                    stat_de_rms = float(match.group(3))
                    stat = (stat_dE_max, stat_dE_avg, stat_de_rms)
                continue
            dE = float(m.group("dE"))
            patch_id = m.group("patch_id")
            error = parse_triplet(m.group("error"))
            color_act = parse_triplet(m.group("color_act"))
            color_ref = parse_triplet(m.group("color_ref"))
            data.append(Patch(patch_id, dE, error, color_act, color_ref))
    if len(data) > 0: return data, stat
    else: return None

def main():
    #parse call arguments
    parser = argparse.ArgumentParser(description=f"Converter from ArgyllCMS:profcheck txt report to xlsx for IT8.7/2 target, v.{_module_date:%Y-%m-%d} by {_module_designer}.")
    parser.add_argument("input", type=Path, help="profcheck report file path")
    parser.add_argument("output", type=Path, nargs='?', help="xlsx file path (optional)")
    parser.add_argument('--nopatchfill', action='store_true', help="Don't fill cell with patch reference color.")
    parser.add_argument('--nocolorgrade', action='store_true', help="Disable color gradations for dE values.")
    args = parser.parse_args()
    if args.output is None:
        args.output = args.input.with_suffix('.xlsx')

    #read profcheck report
    data, stat = read_report(args.input)

    #create xlsx
    workbook = openpyxl.Workbook()

    #IT8.7/2 table overlay ----------------------------------------------------------------------
    worksheet = workbook.active
    worksheet.title = "IT8.7-2"

    #write table headers
    it872_header_col_num(worksheet, 1)
    it872_header_col_num(worksheet, 14)
    it872_header_row_letter(worksheet, 1)
    it872_header_row_letter(worksheet, 24)

    #write patch data into table
    for patch in data:
        cell_row, cell_col = patch_to_cell(patch.id)
        cell = worksheet.cell(row=cell_row, column=cell_col)
        cell.value = patch.dE
        cell.number_format = style_it872_value_numformat
        cell.alignment = style_it872_value_alignment
        #fill cells with patch color
        if args.nopatchfill:
            cell.fill = style_it872_value_fill
        else:
            fill_rgb = pcs_to_rgb(patch.reference)
            fill_hex = "{:02X}{:02X}{:02X}".format(*fill_rgb)
            cell.fill = PatternFill(start_color=fill_hex, end_color=fill_hex, fill_type="solid")
        #color values
        if args.nocolorgrade:
            cell.font = style_it872_value_font
        else:
            font_color = style_it872_value_font.color
            for grade in reversed(dE_gradations):
                if patch.dE >= grade[0]:
                    font_color = grade[1]
                    break
            cell.font = Font(**{**style_it872_value_font.__dict__, "color": font_color})            

    #write stats
    cell_row = 17
    cell_col = 1
    # max
    cell = worksheet.cell(row=cell_row, column=cell_col, value="max:")
    cell.font = style_stat_header_font
    cell.fill = style_stat_header_fill
    cell.alignment = style_stat_header_alignment
    cell = worksheet.cell(row=cell_row, column=cell_col+1, value=stat[0])
    cell.number_format = style_stat_value_numformat
    cell.font = style_stat_value_font
    cell.fill = style_stat_value_fill
    cell.alignment = style_stat_value_alignment
    # avg
    cell = worksheet.cell(row=cell_row+1, column=cell_col, value="avg:")
    cell.font = style_stat_header_font
    cell.fill = style_stat_header_fill
    cell.alignment = style_stat_header_alignment
    cell = worksheet.cell(row=cell_row+1, column=cell_col+1, value=stat[1])
    cell.number_format = style_stat_value_numformat
    cell.font = style_stat_value_font
    cell.fill = style_stat_value_fill
    cell.alignment = style_stat_value_alignment
    # rms
    cell = worksheet.cell(row=cell_row+2, column=cell_col, value="rms:")
    cell.font = style_stat_header_font
    cell.fill = style_stat_header_fill
    cell.alignment = style_stat_header_alignment
    cell = worksheet.cell(row=cell_row+2, column=cell_col+1, value=stat[2])
    cell.number_format = style_stat_value_numformat
    cell.font = style_stat_value_font
    cell.fill = style_stat_value_fill
    cell.alignment = style_stat_value_alignment

    #Grades --------------------------------------------------------------------------------------
    workbook.create_sheet(title="Grades")
    worksheet = workbook.worksheets[1]

    #create addidtional dummy grade to ease condition check
    gradations = dE_gradations + ((375.0, "000000"), )

    for i in range(len(gradations) - 1):
        cell_row = i + 1
        cell_col = 1
        cell = worksheet.cell(row=cell_row, column=cell_col)
        if i == len(gradations) - 2:
            cell.value = f"\u2265 {gradations[i][0]}"
        else:
            cell.value = f"{gradations[i][0]}\u2026{gradations[i+1][0]}"
        if args.nocolorgrade:
            cell.fill = style_grade_header_fill
        else:
            cell.fill = PatternFill(start_color=gradations[i][1], end_color=gradations[i][1], fill_type="solid")

        cell.font = style_grade_header_font
        cell.alignment = style_grade_header_alignment
        worksheet.row_dimensions[cell_row].height = style_grade_cell_size[1]
        for patch in data:
            if gradations[i][0] <= patch.dE < gradations[i+1][0]:
                cell_col += 1
                worksheet.column_dimensions[get_column_letter(cell_col)].width = style_grade_cell_size[0]
                cell = worksheet.cell(row=cell_row, column=cell_col)
                cell.value = patch.id
                cell.font = style_grade_value_font
                cell.alignment = style_grade_value_alignment
                if args.nopatchfill:
                    cell.fill = style_grade_value_fill
                else:
                    fill_rgb = pcs_to_rgb(patch.reference)
                    fill_hex = "{:02X}{:02X}{:02X}".format(*fill_rgb)
                    cell.fill = PatternFill(start_color=fill_hex, end_color=fill_hex, fill_type="solid")
        cell_row += 1

    workbook.save(args.output)
    print(f"Done. xlsx file saved as {args.output}")

if __name__ == "__main__":
    main()